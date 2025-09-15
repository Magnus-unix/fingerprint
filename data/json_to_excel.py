import json
import pandas as pd
import os,sys
os.chdir(sys.path[0])
from openpyxl import load_workbook
from openpyxl.styles import Font
def json_to_excel(json_file, excel_file):
    # 读取 JSON 文件
    with open(json_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    rows = []
    for record in data:
        row = {}
        try:
            rid = int(record.get("id", 0))  # 强制转成整数
        except ValueError:
            continue
        if rid < 326:
            continue

        # 基础字段
        row["username"] = record.get("username")
        row["timestamp"] = record.get("timestamp")
        row["ip"] = record.get("ip")
        #row["cookie"] = record.get("cookie")
        row["url"] = record.get("url")
        row["startTime"] = record.get("startTime")
        row["endTime"] = record.get("endTime")
        
        if row["ip"] == None  and row["url"] == None:
            continue
        # 展开 level1
        for k, v in record.get("level1Signals", {}).items():
            row[f"level1_{k}"] = v

        # 展开 level2
        for k, v in record.get("level2Signals", {}).items():
            row[f"level2_{k}"] = v

        # 展开 level3
        for k, v in record.get("level3Signals", {}).items():
            row[f"level3_{k}"] = v

        rows.append(row)

    # 转换成 DataFrame
    df = pd.DataFrame(rows)
    df.to_excel(excel_file, index=False)

    # 统一设置字体为 Times New Roman
    wb = load_workbook(excel_file)
    ws = wb.active
    font = Font(name="Times New Roman")
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
    wb.save(excel_file)

    print(f"数据已成功导出到 {excel_file} (字体: Times New Roman)")

# 使用示例
if __name__ == "__main__":
    json_to_excel("parsed_records_flat.json", "output.xlsx")
